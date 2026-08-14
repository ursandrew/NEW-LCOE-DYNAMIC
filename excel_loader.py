"""
excel_loader.py — Builds DetailedCostConfig objects directly from Solar_3.xlsm
so category structure and anchor values stay traceable to the source workbook.

This is also the validation entry point: run against the workbook's own
computation table to confirm the engine reproduces its numbers exactly.
"""

import openpyxl
from cost_engine import CostItem, CostCategory, DetailedCostConfig


def _extract_category(ws, start_row: int, n_anchor_cols: int = 6) -> list:
    items = []
    r = start_row
    while True:
        name = ws.cell(row=r, column=2).value
        uom = ws.cell(row=r, column=3).value
        if name is None or 'Subtotal' in str(name):
            break
        values = [ws.cell(row=r, column=c).value for c in range(4, 4 + n_anchor_cols)]
        items.append(CostItem(name=name, uom=uom, values=values))
        r += 1
    return items


def load_solar_config(xlsm_path: str) -> DetailedCostConfig:
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)
    ws = wb['Solar']
    anchors = [ws.cell(row=17, column=c).value for c in range(4, 10)]

    civil = _extract_category(ws, 21)
    mechanical = _extract_category(ws, 30)
    electrical = _extract_category(ws, 37)
    project_mgmt = _extract_category(ws, 49)
    misc = _extract_category(ws, 56)
    opex = _extract_category(ws, 69)

    return DetailedCostConfig(
        anchors_mw=anchors,
        capex_unit_basis='$/Wp',
        capex_categories=[
            # Verified against Solar_3.xlsm computation table (row 85, cap=50 MWp):
            # Base EPC = Civil + Electrical only. Mechanical and Misc are
            # EXCLUDED from Base EPC Cost (matches Excel Flag #3).
            CostCategory('Civil & Construction', civil, counts_toward_base_epc=True),
            CostCategory('Mechanical', mechanical, counts_toward_base_epc=False),
            CostCategory('Electrical', electrical, counts_toward_base_epc=True),
            CostCategory('Miscellaneous', misc, counts_toward_base_epc=False),
            CostCategory('Project Management', project_mgmt),  # % items resolved last
        ],
        opex_items=opex,
    )


def load_wind_config(xlsm_path: str) -> DetailedCostConfig:
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)
    ws = wb['Wind']
    anchors = [ws.cell(row=17, column=c).value for c in range(4, 10)]

    turbine = _extract_category(ws, 21)     # A. Mechanical (Turbine Supply & Installation)
    civil_bop = _extract_category(ws, 28)   # B. Civil (Balance of Plant)
    electrical_grid = _extract_category(ws, 35)  # C. Electrical (Grid Connection)
    project_mgmt = _extract_category(ws, 41)     # D. Project Mgmt & Soft Costs (% items)
    misc = _extract_category(ws, 48)             # E. Miscellaneous
    opex = _extract_category(ws, 61)

    return DetailedCostConfig(
        anchors_mw=anchors,
        capex_unit_basis='$/kW',
        capex_categories=[
            # Verified against Wind sheet summary reference (EPC/PM Fee total =
            # 3,987,000 = 3% x 132,900,000 = Turbine+BOP+Grid, excl. Misc).
            # Base EPC here excludes ONLY Misc — different rule than Solar,
            # confirmed rather than assumed (see Excel Flag #3).
            CostCategory('Turbine Supply & Installation', turbine, counts_toward_base_epc=True),
            CostCategory('Balance of Plant (Civil)', civil_bop, counts_toward_base_epc=True),
            CostCategory('Grid Connection & Substation', electrical_grid, counts_toward_base_epc=True),
            CostCategory('Miscellaneous', misc, counts_toward_base_epc=False),
            CostCategory('Project Management & Soft Costs', project_mgmt),
        ],
        opex_items=opex,
    )


if __name__ == '__main__':
    import sys
    from cost_engine import compute_detailed_capex_opex

    xlsm_path = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/uploads/Solar_3.xlsm'

    print("=" * 70)
    print("VALIDATION: Solar — engine output vs. Excel computation table")
    print("=" * 70)
    solar_cfg = load_solar_config(xlsm_path)

    # Ground truth pulled directly from the Excel's own computation table
    # (COMPUTATION TABLE rows 85+, columns: Civil=10, Mech=16, Elec=27,
    #  ProjMgmt=33, Misc=39, OPEX_Y1=50, GrandCAPEX=51)
    ground_truth = {
        50:  {'grand_capex': 25366171.3,  'opex_y1': 917394.0556},
        60:  {'grand_capex': 30398101.3,  'opex_y1': 1077777.2156},
        70:  {'grand_capex': 35264831.3,  'opex_y1': 1236177.9755999998},
        100: {'grand_capex': 49538271.79, 'opex_y1': 1715459.26148},
    }

    for cap, truth in ground_truth.items():
        result = compute_detailed_capex_opex(solar_cfg, cap)
        capex_diff = result['capital'] - truth['grand_capex']
        opex_diff = result['om_annual'] - truth['opex_y1']
        status = "OK" if abs(capex_diff) < 1 and abs(opex_diff) < 1 else "MISMATCH"
        print(f"\nCapacity {cap} MWp: [{status}]")
        print(f"  CAPEX  engine={result['capital']:,.2f}  excel={truth['grand_capex']:,.2f}  diff={capex_diff:,.4f}")
        print(f"  OPEX   engine={result['om_annual']:,.2f}  excel={truth['opex_y1']:,.2f}  diff={opex_diff:,.4f}")

    print("\n" + "=" * 70)
    print("VALIDATION: Wind config load check")
    print("=" * 70)
    wind_cfg = load_wind_config(xlsm_path)
    for cap in [100, 200, 500]:
        result = compute_detailed_capex_opex(wind_cfg, cap)
        print(f"Capacity {cap} MW -> CAPEX=${result['capital']:,.2f}  OPEX(Y1)=${result['om_annual']:,.2f}")
